import openpyxl

#load up a workbook
newwkbk = openpyxl.Workbook()
fwkbk = openpyxl.load_workbook("sample.xlsx")

#select a sheet
#all workbooks have at least one sheet
mainsheet = fwkbk.active	#loads first sheet, only works in later versions
fwkbk.create_sheet()	#insert sheet at end
fwkbk.create_sheet("Third Sheet")
newwkbk.create_sheet(0)	#insert at first position

#display titles of sheet
print fwkbk.get_sheet_names()	#returns list of sheet names

#change sheet name
fwkbk.get_sheet_by_name('Sheet').title = "Other Sheet"
print fwkbk.get_sheet_names()

fwkbk.worksheets[2].title = "Sheet 3"
print fwkbk.get_sheet_names()

#sheets are iterable
for sheet in fwkbk: print sheet.title

#accessing cells
#can access cells as keys
print mainsheet['C3']		#cell object
print mainsheet['C3'].value
mainsheet['C3'] = -15
print mainsheet['C3'].value
mainsheet['C3'].value = -10
print mainsheet['C3'].value
#print mainsheet.cell('B4').value	#deprecated
print mainsheet.cell(row = 3, column = 3).value

#slicing
for i in mainsheet['A1':'D3']: 		#returns tuple of tuples
	for j in i:
		print j.value

#iter_rows()
#iter_cols()

#to iterate through all rows or columns
#return generators
for i in mainsheet.rows:
	i[2].value *= 2
#there is also a .columns

#writing to file
fwkbk.save("newworkbook.xlsx")



#for dataframe
import pandas as pd
#load to pandas dataframe
df = pd.read_excel('sample.xlsx', sheetname = "Money Lover Report")
df = df.set_index("NO.")		#make our index equal to the id from our report
print df

#write out to new file
writer = pd.ExcelWriter('dataframe.xlsx', engine='xlsxwriter')
df.to_excel(writer,'Money Lover Report')	#ExcelWriter object and sheet name as arguments
